import os
import io
import re
import string
import hashlib
from datetime import datetime
import pandas as pd
import plotly.express as px
import streamlit as st
from dotenv import load_dotenv

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

import sqlalchemy as sa
from sqlalchemy import text

from openai import AzureOpenAI

# -------------------------
# Load environment
# -------------------------
load_dotenv()

try:
    AZURE_OPENAI_ENDPOINT = st.secrets["AZURE_OPENAI_ENDPOINT"]
    AZURE_OPENAI_API_KEY = st.secrets["AZURE_OPENAI_API_KEY"]
    AZURE_OPENAI_DEPLOYMENT = st.secrets["AZURE_OPENAI_DEPLOYMENT"]
    DATABASE = st.secrets["DATABASE"]
except Exception:
    AZURE_OPENAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT")
    AZURE_OPENAI_API_KEY = os.getenv("AZURE_OPENAI_API_KEY")
    AZURE_OPENAI_DEPLOYMENT = os.getenv("AZURE_OPENAI_DEPLOYMENT")
    DATABASE = os.getenv("DATABASE")

if not AZURE_OPENAI_API_KEY or not AZURE_OPENAI_ENDPOINT or not AZURE_OPENAI_DEPLOYMENT:
    st.error("Missing Azure OpenAI credentials. Set AZURE_OPENAI_ENDPOINT, AZURE_OPENAI_API_KEY, AZURE_OPENAI_DEPLOYMENT.")
    st.stop()

DATABASE_URL = (DATABASE or "").strip()

MASTER_XLSX = "genai_job_impact_master.xlsx"
ALL_JOBS_SHEET = "All Jobs"
SYNTHESIS_SHEET = "Synthesis"

# Batch configuration
BATCH_SIZE = 500

st.set_page_config(page_title="GenAI Job Impact Analyst", layout="wide")
st.title("💼 GenAI Job Impact Analyst")

# -------------------------
# Initialize Azure OpenAI client
# -------------------------
client = AzureOpenAI(
    api_key=AZURE_OPENAI_API_KEY,
    api_version="2024-06-01",
    azure_endpoint=AZURE_OPENAI_ENDPOINT,
)

# -------------------------
# Helper functions (verbose)
# -------------------------
def normalize_task(task: str) -> str:
    if not isinstance(task, str):
        task = "" if pd.isna(task) else str(task)
    t = task.lower().strip()
    t = re.sub(r'\s+', ' ', t)
    table = str.maketrans('', '', string.punctuation)
    t = t.translate(table)
    return t

def extract_roles_from_text(raw_text: str) -> list[str]:
    blocks = re.split(r'\n\s*---\s*\n', raw_text.strip(), flags=re.MULTILINE)
    return [b.strip() for b in blocks if b.strip()]

def role_name_from_jobdesc(jd: str, index: int) -> str:
    m = re.search(r'(?i)^ *Job Title:\s*(.+)$', jd, flags=re.MULTILINE)
    if m:
        name = m.group(1).strip()
    else:
        first = next((ln.strip() for ln in jd.splitlines() if ln.strip()), f"Job_{index+1}")
        name = first.split("|")[0].split(" - ")[0].strip()
    name = re.sub(r'[\[\]\*\?/\\:]', "", name)[:120]
    return name or f"Job_{index+1}"

def parse_markdown_table(md_text: str) -> pd.DataFrame:
    lines = [ln.rstrip() for ln in md_text.splitlines()]
    table_lines = [ln for ln in lines if "|" in ln]
    if not table_lines:
        return pd.DataFrame()
    sep_pat = r'^\s*\|?\s*[-:]+(?:\s*\|\s*[-:]+)*\s*\|?\s*$'
    table_lines = [ln for ln in table_lines if not re.match(sep_pat, ln)]
    if not table_lines:
        return pd.DataFrame()
    rows = []
    for ln in table_lines:
        parts = [cell.strip() for cell in ln.strip().strip("|").split("|")]
        rows.append(parts)
    header = rows[0]
    data = rows[1:] if len(rows) > 1 else []
    while header and header[-1] == "":
        header = header[:-1]
        data = [r[:-1] for r in data]
    try:
        df = pd.DataFrame(data, columns=header)
    except Exception:
        max_cols = max(len(r) for r in rows)
        cols = [f"col_{i}" for i in range(max_cols)]
        df = pd.DataFrame([r + [""]*(max_cols-len(r)) for r in rows[1:]], columns=cols)
    df = df.loc[:, ~(df.columns.str.strip() == "")]
    return df

def split_table_and_synthesis(text: str) -> tuple[str, str]:
    parts = text.split("Synthesis:")
    if len(parts) == 2:
        return parts[0], parts[1].strip()
    parts = re.split(r'(?i)Synthèse\s*:', text)
    if len(parts) == 2:
        return parts[0], parts[1].strip()
    return text, ""

# Excel helpers
def load_master_excel() -> tuple[pd.DataFrame, pd.DataFrame]:
    if not os.path.exists(MASTER_XLSX):
        return pd.DataFrame(), pd.DataFrame(columns=["Job Title", "Synthesis", "Run ID", "JD Hash"])
    try:
        xl = pd.ExcelFile(MASTER_XLSX)
        all_jobs = pd.read_excel(xl, sheet_name=ALL_JOBS_SHEET) if ALL_JOBS_SHEET in xl.sheet_names else pd.DataFrame()
        syn = pd.read_excel(xl, sheet_name=SYNTHESIS_SHEET) if SYNTHESIS_SHEET in xl.sheet_names else pd.DataFrame(columns=["Job Title", "Synthesis", "Run ID", "JD Hash"])
        return all_jobs, syn
    except Exception as e:
        st.warning(f"Failed to read master Excel: {e}")
        return pd.DataFrame(), pd.DataFrame(columns=["Job Title", "Synthesis", "Run ID", "JD Hash"])

def write_master_excel(all_jobs: pd.DataFrame, synthesis: pd.DataFrame) -> io.BytesIO:
    if "Job Title" in all_jobs.columns:
        cols = all_jobs.columns.tolist()
        cols = ["Job Title"] + [c for c in cols if c != "Job Title"]
        all_jobs = all_jobs[cols]
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        all_jobs.to_excel(writer, sheet_name=ALL_JOBS_SHEET, index=False)
        synthesis.to_excel(writer, sheet_name=SYNTHESIS_SHEET, index=False)
    wb = load_workbook(MASTER_XLSX)
    if ALL_JOBS_SHEET in wb.sheetnames:
        ws = wb[ALL_JOBS_SHEET]
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for c in col:
                c.alignment = Alignment(wrap_text=True, vertical="top")
                if c.value:
                    l = len(str(c.value))
                    if l > max_len:
                        max_len = l
            ws.column_dimensions[col_letter].width = min(max_len + 2, 80)
    wb.save(MASTER_XLSX)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# Postgres helpers (robust upsert)
def get_engine() -> sa.Engine:
    return sa.create_engine(DATABASE_URL, pool_pre_ping=True, future=True)

def ensure_sql_schema(engine: sa.Engine):
    with engine.begin() as conn:
        conn.execute(text("""
        CREATE TABLE IF NOT EXISTS all_jobs (
            id SERIAL PRIMARY KEY,
            job_title TEXT,
            task TEXT,
            time_allocation TEXT,
            ai_impact_score TEXT,
            impact_explanation TEXT,
            task_transformation TEXT,
            tooling_nature TEXT,
            job_category TEXT,
            Automation_Solution TEXT,
            AI_Automation_Complexity TEXT,
            Upskilling_Suggestion TEXT,
            Task_Category TEXT,
            run_id TEXT,
            jd_hash TEXT,
            task_norm TEXT,
            CONSTRAINT uq_job_task UNIQUE (job_title, task_norm)
        );
        """))
        conn.execute(text("""
        CREATE TABLE IF NOT EXISTS synthesis (
            id SERIAL PRIMARY KEY,
            job_title TEXT,
            synthesis TEXT,
            run_id TEXT,
            jd_hash TEXT,
            CONSTRAINT uq_job_synthesis UNIQUE (job_title, jd_hash)
        );
        """))

def upsert_all_jobs_sql(engine: sa.Engine, df: pd.DataFrame):
    if df is None or df.empty:
        return
    expected_cols = [
        "Job Title",
        "Task",
        "Time allocation %",
        "AI Impact Score (0–100)",
        "Impact Explanation",
        "Task Transformation %",
        "Tooling nature % generic vs specific",
        "Job Category",
        "Automation Solution",
        "AI Automation Complexity",
        "Upskilling Suggestion",
        "Task Category",
        "Run ID",
        "JD Hash",
        "task_norm"
    ]
    col_map = {}
    for c in df.columns:
        cc = c.strip()
        lower = cc.lower()
        if lower in ["task", "tasks"]:
            col_map[c] = "Task"
        elif "job category" in lower:
            col_map[c] = "Job Category"
        elif "time" in lower and "alloc" in lower:
            col_map[c] = "Time allocation %"
        elif "ai impact" in lower or "impact score" in lower:
            col_map[c] = "AI Impact Score (0–100)"
        elif "impact explanation" in lower or ("explanation" in lower and "impact" in lower):
            col_map[c] = "Impact Explanation"
        elif "task transformation" in lower or "transformation" in lower:
            col_map[c] = "Task Transformation %"
        elif "tooling" in lower:
            col_map[c] = "Tooling nature % generic vs specific"
        elif "automation solution" in lower or ("solution" in lower and "automation" in lower):
            col_map[c] = "Automation Solution"
        elif "ai automation complexity" in lower or ("complexity" in lower and "automation" in lower):
            col_map[c] = "AI Automation Complexity"
        elif "upskilling suggestion" in lower or ("upskilling" in lower and "suggestion" in lower):
            col_map[c] = "Upskilling Suggestion"
        elif "task category" in lower:
            col_map[c] = "Task Category"
        elif "job title" in lower or lower == "title":
            col_map[c] = "Job Title"
        elif lower in ["run id", "run_id", "runid"]:
            col_map[c] = "Run ID"
        elif lower in ["jd hash", "jd_hash", "jdhash"]:
            col_map[c] = "JD Hash"
        elif lower == "task_norm":
            col_map[c] = "task_norm"
    if col_map:
        df = df.rename(columns=col_map)
    for col in expected_cols:
        if col not in df.columns:
            df[col] = None
    if "task_norm" not in df.columns or df["task_norm"].isnull().all():
        if "Task" in df.columns:
            df["task_norm"] = df["Task"].apply(normalize_task)
        else:
            df["task_norm"] = ""
    out = df[expected_cols].copy().where(pd.notnull(df), None)
    out = out.drop_duplicates(subset=["Job Title", "task_norm"], keep="last").reset_index(drop=True)
    rows = out.to_dict(orient="records")
    insert_stmt = text("""
        INSERT INTO all_jobs
        (job_title, task, time_allocation, ai_impact_score, impact_explanation,
         task_transformation, tooling_nature, job_category, Automation_Solution, AI_AutomATION_Complexity, Upskilling_Suggestion, Task_Category, run_id, jd_hash, task_norm)
        VALUES (:job_title, :task, :time_allocation, :ai_impact_score, :impact_explanation,
                :task_transformation, :tooling_nature, :job_category, :Automation_Solution, :AI_AutomATION_Complexity,:Upskilling_Suggestion, :Task_Category, :run_id, :jd_hash, :task_norm)
        ON CONFLICT (job_title, task_norm) DO UPDATE SET
            time_allocation = EXCLUDED.time_allocation,
            ai_impact_score = EXCLUDED.ai_impact_score,
            impact_explanation = EXCLUDED.impact_explanation,
            task_transformation = EXCLUDED.task_transformation,
            tooling_nature = EXCLUDED.tooling_nature,
            job_category = EXCLUDED.job_category,
            Automation_Solution = Excluded.Automation_Solution,
            AI_AutomATION_Complexity = Excluded.AI_AutomATION_Complexity,
            Upskilling_Suggestion = Excluded.Upskilling_Suggestion,
            Task_Category = Excluded.Task_Category,
            run_id = EXCLUDED.run_id,
            jd_hash = EXCLUDED.jd_hash;
    """)
    stmt_rows = []
    for r in rows:
        stmt_rows.append({
            "job_title": r.get("Job Title"),
            "task": r.get("Task"),
            "time_allocation": r.get("Time allocation %"),
            "ai_impact_score": r.get("AI Impact Score (0–100)"),
            "impact_explanation": r.get("Impact Explanation"),
            "task_transformation": r.get("Task Transformation %"),
            "tooling_nature": r.get("Tooling nature % generic vs specific"),
            "job_category": r.get("Job Category"),
            "Automation_Solution": r.get("Automation Solution"),
            "AI_AutomATION_Complexity": r.get("AI Automation Complexity"),
            "Upskilling_Suggestion": r.get("Upskilling Suggestion"),
            "Task_Category": r.get("Task Category"),
            "run_id": r.get("Run ID"),
            "jd_hash": r.get("JD Hash"),
            "task_norm": r.get("task_norm")
        })
    with engine.begin() as conn:
        if stmt_rows:
            conn.execute(insert_stmt, stmt_rows)

def append_synthesis_sql(engine: sa.Engine, syn_rows: list):
    if not syn_rows:
        return
    insert_stmt = text("""
        INSERT INTO synthesis (job_title, synthesis, run_id, jd_hash)
        VALUES (:job_title, :synthesis, :run_id, :jd_hash)
        ON CONFLICT (job_title, jd_hash) DO NOTHING;
    """)
    with engine.begin() as conn:
        conn.execute(insert_stmt, syn_rows)

# -------------------------
# Session-state buffers
# -------------------------
if "new_reports" not in st.session_state:
    st.session_state["new_reports"] = {}
if "new_synthesis" not in st.session_state:
    st.session_state["new_synthesis"] = {}
if "new_jd_text" not in st.session_state:
    st.session_state["new_jd_text"] = {}
if "all_job_descriptions" not in st.session_state:
    st.session_state["all_job_descriptions"] = []
if "current_batch" not in st.session_state:
    st.session_state["current_batch"] = 0
if "total_batches" not in st.session_state:
    st.session_state["total_batches"] = 0
if "auto_run_batches" not in st.session_state:
    st.session_state["auto_run_batches"] = False

# -------------------------
# Sidebar / upload UI
# -------------------------
st.sidebar.header("Upload or Write Job Description(s)")
uploaded_file = st.sidebar.file_uploader("Upload job descriptions (.txt or .csv)", type=["txt", "csv"])
job_text = st.sidebar.text_area("Or paste a single job description here")
st.sidebar.caption(
    "💡 Multiple roles? In .txt, separate with a line containing only `---`. "
    "In .csv, provide one job description per row under a column named 'JobDescription'."
)

# -------------------------
# Load job descriptions button
# -------------------------
if st.sidebar.button("📂 Load Job Descriptions"):
    job_descriptions = []
    if uploaded_file is not None:
        if uploaded_file.name.endswith(".txt"):
            raw_text = uploaded_file.read().decode("utf-8")
            job_descriptions = extract_roles_from_text(raw_text)
        elif uploaded_file.name.endswith(".csv"):
            try:
                df_csv = pd.read_csv(uploaded_file)
            except Exception as e:
                st.error(f"Could not read CSV: {e}")
                st.stop()
            # Case-insensitive column handling for CSV
            cols_map = {c.lower().strip(): c for c in df_csv.columns}
            chosen_col = None
            for candidate in ("jobdescription", "job_description", "job description", "description", "jobdesc", "job_desc"):
                if candidate in cols_map:
                    chosen_col = cols_map[candidate]
                    break
            if chosen_col:
                job_descriptions = df_csv[chosen_col].dropna().astype(str).tolist()
            else:
                st.error("CSV must contain a column named 'JobDescription' or similar (job_description, description).")
                st.stop()
        else:
            st.error("Unsupported file type. Upload .txt or .csv")
            st.stop()
    elif job_text and job_text.strip():
        job_descriptions = [job_text.strip()]
    else:
        st.error("Please upload or paste at least one job description.")
        st.stop()
    
    st.session_state["all_job_descriptions"] = job_descriptions
    st.session_state["total_batches"] = (len(job_descriptions) + BATCH_SIZE - 1) // BATCH_SIZE
    st.session_state["current_batch"] = 0
    st.success(f"✅ Loaded {len(job_descriptions)} job descriptions. Total batches: {st.session_state['total_batches']}")

# -------------------------
# Batch Processing Controls (with Power BI + Auto-Run)
# -------------------------
jobs = st.session_state.get("all_job_descriptions", [])
if jobs:
    total_jobs = len(jobs)
    batch_size_local = BATCH_SIZE

    # Ensure batch counters exist
    if "current_batch" not in st.session_state:
        st.session_state["current_batch"] = 0

    total_batches = (total_jobs + batch_size_local - 1) // batch_size_local
    st.session_state["total_batches"] = total_batches

    st.sidebar.divider()
    st.sidebar.subheader("📦 Batch Processing")

    st.sidebar.info(f"Total jobs: {total_jobs} | Batch size: {batch_size_local}")

    # -------------------------
    # Navigation Buttons
    # -------------------------
    col_b1, col_b2 = st.sidebar.columns(2)

    with col_b1:
        st.button(
            "⬅️ Previous",
            disabled=st.session_state["current_batch"] <= 0,
            on_click=lambda: st.session_state.__setitem__(
                "current_batch",
                max(0, st.session_state["current_batch"] - 1)
            )
        )

    with col_b2:
        st.button(
            "Next ➡️",
            disabled=st.session_state["current_batch"] >= total_batches - 1,
            on_click=lambda: st.session_state.__setitem__(
                "current_batch",
                min(total_batches - 1, st.session_state["current_batch"] + 1)
            )
        )

    # -------------------------
    # Select Batch Dropdown
    # -------------------------
    selected = st.sidebar.selectbox(
        "Jump to batch:",
        list(range(total_batches)),
        index=st.session_state["current_batch"],
        format_func=lambda x: f"Batch {x+1} (Jobs {x*batch_size_local+1}-{min((x+1)*batch_size_local, total_jobs)})",
        key="batch_selector"
    )

    if selected != st.session_state["current_batch"]:
        st.session_state["current_batch"] = selected
        st.rerun()

    st.sidebar.markdown(f"### ▶ Current Batch: **{st.session_state['current_batch'] + 1} / {total_batches}**")

    # -------------------------
    # Generate Batch Button + Power BI Button (Main Area)
    # -------------------------
    col_main1, col_main2 = st.columns([1, 1])

    with col_main1:
        st.button(
            "🚀 Generate Report for This Batch",
            type="primary",
            on_click=lambda: st.session_state.__setitem__("trigger_process_batch", True)
        )

    with col_main2:
        powerbi_url = "https://app.powerbi.com/view?r=eyJrIjoiMDFhMGVlOGItOTY5MC00ZTRhLWI5ZTEtNmMwNDQxNTUzNTNmIiwidCI6IjA3NmEzOTkyLTA0ZjgtNDcwMC05ODQ0LTA4YzM3NDc3NzdlZiJ9"
        st.markdown(
            f"""
            <a href="{powerbi_url}" target="_blank">
                <button style="background-color:#0078D4; color:white; padding:0.6em 1.2em; border:none; border-radius:8px; cursor:pointer;">
                    📊 Open Dashboard
                </button>
            </a>
            """,
            unsafe_allow_html=True
        )

    # -------------------------
    # Sidebar version of Generate Report (duplicate)
    # -------------------------
    if st.sidebar.button("🚀 Generate Report for This Batch", type="primary"):
        st.session_state["trigger_process_batch"] = True

    # -------------------------
    # Auto-Run ALL Batches (Generate all, Commit per-batch)
    # -------------------------
    if st.sidebar.button("⚡ Auto-Run All Batches (Generate + Commit per batch)"):
        st.session_state["auto_run_batches"] = True

else:
    st.sidebar.info("No job descriptions loaded yet.")
# -------------------------
# DB connection check (optional)
# -------------------------
engine = None
if DATABASE_URL:
    try:
        engine = get_engine()
        with engine.connect() as c:
            c.execute(sa.text("SELECT 1"))
        st.success("✅ Postgres connection OK (DB commit button will be active).")
        ensure_sql_schema(engine)
    except Exception as e:
        st.warning(f"DB connection failed or not available: {e}. You can still update Excel only.")
        engine = None
else:
    st.info("No DATABASE configured — app will operate in Excel-only mode unless DATABASE is provided.")

# -------------------------
# SYSTEM PROMPT
# -------------------------
SYSTEM_PROMPT = """You are GenAI-Job-Impact-Analyst, an expert designed to evaluate how generative AI can transform work at Club Med. 

Your mission
Input: You will receive either
  - a full text job description, OR
  - just a job title (with little or no detail).

If only a job title is given, infer the typical tasks and responsibilities for that role at Club Med or in the hospitality industry, and continue as if a full description was provided.

Output: Produce a table – one line per task – with the following six columns: 
| Task | Job Category | Time allocation % | AI Impact Score (0–100) | Impact Explanation | Task Transformation % | Tooling nature % generic vs specific | Automation Solution | AI Automation Complexity | Upskilling Suggestion | Task Category |

Task – concise verb-phrase copied, paraphrased, or reasonably inferred from the job title or description. 
Job Category - one of: IT, Marketing, HR, Finance, Operations, Legal, R&D, Customer Service, Other.
Time allocation % – your best estimate of the share of the job's total time this task takes (sum ≈ 100%). 
AI Impact Score – how strongly Gen-AI could affect the task (0 = no impact, 100 = fully automatable/augmented). 
Impact Explanation – 2–3 sentences justifying the chosen score. Write the Impact Explanation only in French.
Task Transformation % – proportion of the task likely to change for the employee (e.g., 70% up-skilling vs 30% pure automation). Always express as two percentages that sum to 100 in the format "XX% up-skilling / YY% automation".
Tooling nature – split the AI tooling you foresee into generic (ChatGPT-like) vs domain-specific (custom models or vertical SaaS). Express as two numbers that sum to 100.
Automation Solution – briefly describe a realistic Gen-AI solution (e.g., "custom GPT-4 powered chatbot", "AI-assisted code generation tool", "AI-driven marketing content generator").
AI Automation Complexity – rate the complexity of building and deploying the AI solution (1 = very simple, 5 = very complex).
Upskilling Suggestion – Suggest one or two key skills the employee should develop to thrive in an AI-augmented version of this task. Write the Upskilling Suggestion only in French.
Task Category – Knowledge Work / Physical Work / Hybrid Work (Knowledge work = computer-based, cognitive tasks like, planning, coordinating, communicating, deciding. Physical Work = tasks whose description clearly implies on-site/manual presence e.g. tasks with keywords like welcome guests, check-in, serving at bar/restaurant, cooking, cleaning, housekeeping, performing in shows, maintenance, driving, lifeguard, etc. Hybrid Work = a mix of both knowledge and physical work).

Procedure
A. If a detailed description is given: scan the description and list every distinct, non-trivial activity. 
B. If only a job title is given: generate a reasonable list of 5–10 core tasks typical for the role in hospitality / Club Med.
C. Estimate Time allocation % first – it anchors later scores. Round to nearest 5%.
D. For each activity, consider whether Gen-AI could draft, summarize, translate, ideate, classify, predict or converse, and estimate the effect.
E. Deliver the table in Markdown, then add a short one-paragraph synthesis highlighting the top three automation opportunities and any human-core tasks that should remain manual.

Formatting rules
Use Markdown. Keep lines reasonably wrapped (~80 chars). Round percentages to nearest 5%. Do not invent tasks that are absent when a detailed JD is provided. Never return an empty output — if input is a title only, infer typical tasks and still return a full table + synthesis.
"""

# -------------------------
# Process single job description
# -------------------------
def process_single_job(jd: str, idx: int, global_idx: int):
    """Process a single job description and return the results"""
    role_name = role_name_from_jobdesc(jd, global_idx)
    user_prompt = f"Here is the job description or job title:\n\n{jd}"

    try:
        resp = client.chat.completions.create(
            model=AZURE_OPENAI_DEPLOYMENT,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": user_prompt},
            ]
        )
        output_text = resp.choices[0].message.content
    except Exception as e:
        st.error(f"OpenAI call failed for {role_name}: {e}")
        output_text = ""

    # Fallback: if the model returned nothing, try an inferred-tasks prompt
    if not output_text or not output_text.strip():
        st.warning(f"No detailed output from model for {role_name}. Using inferred tasks fallback.")
        fallback_prompt = f"Please generate typical tasks for the role '{role_name}' (as used in hospitality/Club Med) and evaluate them following the instructions."
        try:
            resp = client.chat.completions.create(
                model=AZURE_OPENAI_DEPLOYMENT,
                messages=[
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user", "content": fallback_prompt},
                ]
            )
            output_text = resp.choices[0].message.content
        except Exception as e:
            st.error(f"Fallback OpenAI call failed for {role_name}: {e}")
            output_text = ""

    table_text, synthesis_text = split_table_and_synthesis(output_text)
    parsed_df = parse_markdown_table(table_text)

    if parsed_df.empty:
        m = re.search(r"(\|.*\|\s*\n\|[-:\s|]+\|\s*\n(?:\|.*\|\s*\n?)*)", output_text, flags=re.DOTALL)
        if m:
            parsed_df = parse_markdown_table(m.group(0))
        if parsed_df.empty:
            parsed_df = pd.DataFrame({
                "Task": [f"[Model output parse failed — see Report]"],
                "AI Impact Score (0–100)": [None],
                "Job Category": [None],
                "Time allocation %": [None],
                "Impact Explanation": [None],
                "Task Transformation %": [None],
                "Tooling nature % generic vs specific": [None],
                "Automation Solution": [None],
                "AI Automation Complexity": [None],
                "Upskilling Suggestion": [None],
                "Task Category": [None]
            })

    if "Job Title" not in parsed_df.columns:
        parsed_df.insert(0, "Job Title", role_name)
    else:
        parsed_df["Job Title"] = parsed_df["Job Title"].replace("", role_name).fillna(role_name)

    canonical_map = {}
    for col in parsed_df.columns:
        lc = col.strip().lower()
        if lc in ["task", "tasks"]:
            canonical_map[col] = "Task"
        elif "job category" in lc:
            canonical_map[col] = "Job Category"
        elif "time" in lc and "alloc" in lc:
            canonical_map[col] = "Time allocation %"
        elif "ai impact" in lc or "impact score" in lc:
            canonical_map[col] = "AI Impact Score (0–100)"
        elif "impact explanation" in lc or ("explanation" in lc and "impact" in lc):
            canonical_map[col] = "Impact Explanation"
        elif "task transformation" in lc or "transformation" in lc:
            canonical_map[col] = "Task Transformation %"
        elif "tooling" in lc:
            canonical_map[col] = "Tooling nature % generic vs specific"
        elif "automation solution" in lc or ("solution" in lc and "automation" in lc):
            canonical_map[col] = "Automation Solution"
        elif "ai automation complexity" in lc or ("complexity" in lc and "automation" in lc):
            canonical_map[col] = "AI Automation Complexity"
        elif "upskilling suggestion" in lc or ("upskilling" in lc and "suggestion" in lc):
            canonical_map[col] = "Upskilling Suggestion"
        elif "task category" in lc or ("category" in lc and "task" in lc):
            canonical_map[col] = "Task Category"
        elif lc in ["job title", "title"]:
            canonical_map[col] = "Job Title"
    if canonical_map:
        parsed_df = parsed_df.rename(columns=canonical_map)

    for col in [
        "Task",
        "Job Category",
        "Time allocation %",
        "AI Impact Score (0–100)",
        "Impact Explanation",
        "Task Transformation %",
        "Tooling nature % generic vs specific",
        "Automation Solution",
        "AI Automation Complexity",
        "Upskilling Suggestion",
        "Task Category"
    ]:
        if col not in parsed_df.columns:
            parsed_df[col] = None

    run_id = datetime.now().isoformat(timespec="seconds")
    jd_hash = hashlib.sha256(jd.strip().encode("utf-8")).hexdigest()[:12]
    parsed_df["Run ID"] = run_id
    parsed_df["JD Hash"] = jd_hash

    if "Task" in parsed_df.columns:
        parsed_df["task_norm"] = parsed_df["Task"].apply(normalize_task)
        parsed_df = parsed_df.drop_duplicates(subset=["Job Title", "task_norm"], keep="last")
    else:
        parsed_df["task_norm"] = ""

    # Normalize Task Transformation % into "XX% up-skilling / YY% automation"
    def normalize_task_transformation(val):
        if val is None:
            return None
        s = str(val).strip()
        if "%" in s and ("/" in s or "up" in s.lower()):
            return s
        m = re.search(r'(\d{1,3})', s)
        if m:
            num = int(m.group(1))
            if num < 0: num = 0
            if num > 100: num = 100
            other = 100 - num
            return f"{num}% up-skilling / {other}% automation"
        return None

    if "Task Transformation %" in parsed_df.columns:
        parsed_df["Task Transformation %"] = parsed_df["Task Transformation %"].apply(normalize_task_transformation)

    cols = parsed_df.columns.tolist()
    if "Job Title" in cols:
        cols = ["Job Title"] + [c for c in cols if c != "Job Title"]
        parsed_df = parsed_df[cols]

    return role_name, parsed_df, synthesis_text, output_text, jd
# -------------------------
# GENERATE: batch processing (manual trigger)
# -------------------------
if st.session_state.get("trigger_process_batch", False):
    st.session_state["trigger_process_batch"] = False  # reset the trigger

    current_batch = st.session_state.get("current_batch", 0)
    start_idx = current_batch * BATCH_SIZE
    end_idx = min(start_idx + BATCH_SIZE, len(st.session_state.get("all_job_descriptions", [])))

    batch_jobs = st.session_state.get("all_job_descriptions", [])[start_idx:end_idx]

    st.info(f"🔄 Processing Batch {current_batch + 1}/{st.session_state.get('total_batches', 1)} (Jobs {start_idx + 1}-{end_idx})")

    # Progress bar
    progress_bar = st.progress(0)
    status_text = st.empty()

    with st.spinner(f"Analyzing {len(batch_jobs)} job descriptions with Azure OpenAI..."):
        for idx, jd in enumerate(batch_jobs):
            global_idx = start_idx + idx

            # Update progress
            progress = (idx + 1) / len(batch_jobs) if len(batch_jobs) > 0 else 1.0
            progress_bar.progress(progress)
            status_text.text(f"Processing job {idx + 1}/{len(batch_jobs)} (Global: {global_idx + 1}/{len(st.session_state.get('all_job_descriptions', []))})")

            # Process the job
            role_name, parsed_df, synthesis_text, output_text, jd_text = process_single_job(jd, idx, global_idx)

            # Display results
            with st.expander(f"📊 Generated Report — **{role_name}** (Job {global_idx + 1})"):
                if output_text:
                    st.markdown(output_text)
                else:
                    st.markdown("_No output from model._")

            # Store in session state
            st.session_state["new_reports"][role_name] = parsed_df.copy()
            st.session_state["new_synthesis"][role_name] = synthesis_text
            st.session_state["new_jd_text"][role_name] = jd_text

    progress_bar.progress(1.0)
    status_text.text(f"✅ Completed processing batch {current_batch + 1}")
    st.success(f"✅ Successfully processed {len(batch_jobs)} job descriptions in this batch!")

# ------------------------------
# AUTO-RUN ALL BATCHES (Generate each batch, then COMMIT per-batch)
# ------------------------------
if st.session_state.get("auto_run_batches", False):
    # reset trigger right away
    st.session_state["auto_run_batches"] = False

    all_jobs = st.session_state.get("all_job_descriptions", [])
    total_jobs = len(all_jobs)
    total_batches = st.session_state.get("total_batches", (total_jobs + BATCH_SIZE - 1)//BATCH_SIZE)
    st.info(f"⚡ Auto-run started: {total_jobs} jobs across {total_batches} batches.")
    overall_progress = st.progress(0.0)
    overall_status = st.empty()

    failed_batches = []
    processed_total = 0
    batch_summaries = []

    for batch_num in range(total_batches):
        st.session_state["current_batch"] = batch_num
        start_idx = batch_num * BATCH_SIZE
        end_idx = min(start_idx + BATCH_SIZE, total_jobs)
        batch_jobs = all_jobs[start_idx:end_idx]

        overall_status.write(f"🔄 Generating Batch {batch_num + 1}/{total_batches} (Jobs {start_idx + 1}-{end_idx})")

        # per-batch progress
        inner_progress = st.progress(0.0)
        job_status = st.empty()

        # store roles generated in this batch so we can commit only them
        batch_role_names = []

        with st.spinner(f"Generating batch {batch_num+1}..."):
            try:
                for i, jd in enumerate(batch_jobs):
                    global_idx = start_idx + i

                    role_name, parsed_df, synthesis_text, output_text, jd_text = process_single_job(jd, i, global_idx)

                    # Display output (same as manual)
                    with st.expander(f"📊 Generated Report — **{role_name}** (Global Job {global_idx+1})"):
                        if output_text:
                            st.markdown(output_text)
                        else:
                            st.warning("No output from model.")

                    # Store in global buffers (so commit logic can find them)
                    st.session_state["new_reports"][role_name] = parsed_df.copy()
                    st.session_state["new_synthesis"][role_name] = synthesis_text
                    st.session_state["new_jd_text"][role_name] = jd_text

                    batch_role_names.append(role_name)
                    processed_total += 1

                    # progress updates
                    pct = (i + 1) / len(batch_jobs)
                    inner_progress.progress(pct)
                    job_status.text(f"Batch {batch_num+1}/{total_batches} — Job {i+1}/{len(batch_jobs)} processed")

            except Exception as e:
                failed_batches.append(batch_num + 1)
                st.error(f"Batch {batch_num+1} failed during generation: {e}")

        # After generation of this batch -> COMMIT this batch to DB (if engine exists), show DB progress
        batch_summary = {"batch": batch_num + 1, "start": start_idx + 1, "end": end_idx, "rows": 0, "synth": 0, "status": "pending"}
        try:
            if engine is None:
                st.warning("No DB configured — skipping DB commit for this batch (Excel-only mode).")
                # Still count rows and synthesis items
                try:
                    batch_dfs = [st.session_state["new_reports"][rn] for rn in batch_role_names if rn in st.session_state["new_reports"]]
                    batch_df = pd.concat(batch_dfs, ignore_index=True) if batch_dfs else pd.DataFrame()
                    batch_summary["rows"] = len(batch_df)
                    batch_summary["synth"] = len([rn for rn in batch_role_names if rn in st.session_state["new_synthesis"]])
                    batch_summary["status"] = "excel-only"
                except Exception:
                    batch_summary["status"] = "excel-only-failed"
            else:
                st.info(f"🗄️ Committing Batch {batch_num+1} to database...")
                batch_db_progress = st.progress(0.0)
                batch_db_status = st.empty()

                # Build DF for this batch only
                batch_db_status.text("Preparing batch data...")
                batch_db_progress.progress(0.05)

                batch_dfs = [st.session_state["new_reports"][rn] for rn in batch_role_names if rn in st.session_state["new_reports"]]
                batch_df = pd.concat(batch_dfs, ignore_index=True) if batch_dfs else pd.DataFrame()

                batch_summary["rows"] = len(batch_df)

                # Clean Task Transformation %
                if "Task Transformation %" in batch_df.columns:
                    def fix_tf(v):
                        if isinstance(v, str) and "%" in v:
                            return v
                        try:
                            m = re.search(r"(\d+)", str(v))
                            n = int(m.group(1)) if m else 50
                        except:
                            n = 50
                        n = max(0, min(100, n))
                        return f"{n}% up-skilling / {100-n}% automation"

                    batch_df["Task Transformation %"] = batch_df["Task Transformation %"].apply(fix_tf)

                # Phase 1 — upsert tasks
                batch_db_status.text("Inserting tasks into database...")
                batch_db_progress.progress(0.35)

                upsert_all_jobs_sql(engine, batch_df)
                batch_db_progress.progress(0.7)
                batch_db_status.text("Tasks inserted successfully.")

                # Phase 2 — append synth
                syn_rows = []
                for rn in batch_role_names:
                    if rn in st.session_state["new_synthesis"]:
                        syn_text = st.session_state["new_synthesis"][rn]
                        jd_text = st.session_state["new_jd_text"].get(rn, "")
                        syn_rows.append({
                            "job_title": rn,
                            "synthesis": syn_text,
                            "run_id": datetime.now().isoformat(timespec="seconds"),
                            "jd_hash": hashlib.sha256(jd_text.strip().encode("utf-8")).hexdigest()[:12]
                        })

                batch_summary["synth"] = len(syn_rows)
                batch_db_status.text("Inserting synthesis rows...")
                batch_db_progress.progress(0.9)

                if syn_rows:
                    append_synthesis_sql(engine, syn_rows)

                batch_db_status.text("Batch commit finalizing...")
                batch_db_progress.progress(1.0)
                batch_db_status.text("Batch commit complete.")
                batch_summary["status"] = "committed"
                st.success(f"✅ Batch {batch_num+1} committed to database ({batch_summary['rows']} rows, {batch_summary['synth']} synth).")

        except Exception as e:
            batch_summary["status"] = f"failed: {e}"
            st.error(f"❌ DB commit failed for batch {batch_num+1}: {e}")
            failed_batches.append(batch_num + 1)

        # Clear only the items that belong to this batch from session_state buffers
        for rn in batch_role_names:
            if rn in st.session_state["new_reports"]:
                del st.session_state["new_reports"][rn]
            if rn in st.session_state["new_synthesis"]:
                del st.session_state["new_synthesis"][rn]
            if rn in st.session_state["new_jd_text"]:
                del st.session_state["new_jd_text"][rn]

        batch_summaries.append(batch_summary)

        # update overall progress
        overall_progress.progress((batch_num + 1) / total_batches)

    # End for all batches
    # Summary of auto-run
    st.info("📝 Auto-run summary")
    summary_df = pd.DataFrame(batch_summaries)
    st.dataframe(summary_df)

    if failed_batches:
        st.warning(f"Some batches failed: {failed_batches}")
    else:
        st.success("🎉 All batches generated and committed successfully.")

    st.balloons()
# -------------------------
# Preview buffered items
# -------------------------

st.divider()
st.subheader("📝 Pending Updates (Buffered; choose how to commit)")

if st.session_state["new_reports"]:
    st.info(f"📦 Total buffered reports: {len(st.session_state['new_reports'])}")
    
    # Show summary
    with st.expander("View Summary of Buffered Jobs", expanded=False):
        buffered_jobs = list(st.session_state["new_reports"].keys())
        for i, job in enumerate(buffered_jobs, 1):
            st.write(f"{i}. {job}")
    
    # Show detailed view (limit to last 5 for performance)
    st.markdown("#### Last 5 Processed Jobs (Detailed View)")
    recent_jobs = list(st.session_state["new_reports"].items())[-5:]
    for role, df in recent_jobs:
        with st.expander(f"📊 {role}"):
            try:
                st.dataframe(df, use_container_width=True)
            except Exception:
                st.write(df.head(20))
            syn = st.session_state["new_synthesis"].get(role, "")
            if syn:
                st.markdown("**Synthesis (preview):**")
                st.markdown(syn if len(syn) < 1000 else syn[:1000] + "...")
else:
    st.info("No buffered results. Use 'Generate Report' to parse JDs.")

# -------------------------
# Two separate commit buttons
# -------------------------
st.divider()
st.subheader("💾 Commit Options")

col_a, col_b = st.columns(2)

with col_a:
    excel_disabled = not bool(st.session_state["new_reports"])
    if st.button("📊 Update Master Excel (Excel only)", disabled=excel_disabled):
        existing_tasks, existing_syn = load_master_excel()
        try:
            new_tasks = pd.concat(st.session_state["new_reports"].values(), ignore_index=True, sort=False)
        except Exception:
            new_tasks = pd.DataFrame()
            for v in st.session_state["new_reports"].values():
                new_tasks = pd.concat([new_tasks, v], ignore_index=True, sort=False)

        if "Task" in new_tasks.columns:
            new_tasks["task_norm"] = new_tasks["Task"].apply(normalize_task)
        else:
            new_tasks["task_norm"] = ""

        if not existing_tasks.empty and "Task" in existing_tasks.columns:
            existing_tasks["task_norm"] = existing_tasks["Task"].apply(normalize_task)
        else:
            if existing_tasks.empty:
                existing_tasks = pd.DataFrame(columns=new_tasks.columns.tolist())

        if existing_tasks.empty:
            all_tasks = new_tasks.copy()
        else:
            key_cols = ["Job Title", "task_norm"]
            existing_keys = existing_tasks[key_cols].drop_duplicates()
            merged = new_tasks.merge(existing_keys, on=key_cols, how="left", indicator=True)
            to_add = merged[merged["_merge"] == "left_only"].drop(columns=["_merge"])
            all_tasks = pd.concat([existing_tasks, to_add], ignore_index=True, sort=False)

        if "task_norm" in all_tasks.columns:
            cols = [c for c in all_tasks.columns if c != "task_norm"] + ["task_norm"]
            all_tasks = all_tasks[cols]

        new_syn_rows = []
        for role, syn in st.session_state["new_synthesis"].items():
            jd_text = st.session_state["new_jd_text"].get(role, "")
            run_id = datetime.now().isoformat(timespec="seconds")
            jd_hash = hashlib.sha256(jd_text.strip().encode("utf-8")).hexdigest()[:12]
            new_syn_rows.append({
                "Job Title": role,
                "Synthesis": syn,
                "Run ID": run_id,
                "JD Hash": jd_hash
            })
        new_syn_df = pd.DataFrame(new_syn_rows, columns=["Job Title", "Synthesis", "Run ID", "JD Hash"])

        if existing_syn is None or existing_syn.empty:
            all_syn = new_syn_df.copy()
        else:
            all_syn = pd.concat([existing_syn, new_syn_df], ignore_index=True, sort=False)
            if "JD Hash" in all_syn.columns:
                all_syn = all_syn.drop_duplicates(subset=["Job Title", "JD Hash"], keep="last")

        expected_cols_for_excel = [
            "Job Title",
            "Task",
            "Job Category",
            "Time allocation %",
            "AI Impact Score (0–100)",
            "Impact Explanation",
            "Task Transformation %",
            "Tooling nature % generic vs specific",
            "Automation Solution",
            "AI Automation Complexity",
            "Upskilling Suggestion",
            "Task Category",
            "Run ID",
            "JD Hash"
        ]
        for col in expected_cols_for_excel:
            if col not in all_tasks.columns:
                all_tasks[col] = None
        extras = [c for c in all_tasks.columns if c not in expected_cols_for_excel]
        final_cols = expected_cols_for_excel + extras
        all_tasks = all_tasks[final_cols]

        try:
            buf = write_master_excel(all_tasks, all_syn)
            st.success(f"✅ Master Excel updated: {MASTER_XLSX} (database not modified). Total jobs in Excel: {len(all_tasks)}")
            st.download_button(
                label="📥 Download Current Master Excel",
                data=buf,
                file_name=MASTER_XLSX,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"Failed to write master Excel: {e}")

        st.session_state["new_reports"].clear()
        st.session_state["new_synthesis"].clear()
        st.session_state["new_jd_text"].clear()
        st.rerun()

with col_b:
    db_disabled = (not bool(st.session_state["new_reports"])) or (engine is None)
    if st.button("🗄️ Update Database (Postgres only)", disabled=db_disabled):
        if engine is None:
            st.error("No DB engine available. Configure DATABASE to enable DB updates.")
        else:
            try:
                new_tasks = pd.concat(st.session_state["new_reports"].values(), ignore_index=True, sort=False)
            except Exception:
                new_tasks = pd.DataFrame()
                for v in st.session_state["new_reports"].values():
                    new_tasks = pd.concat([new_tasks, v], ignore_index=True, sort=False)

            # Ensure Task Transformation formatting before DB upsert
            def normalize_task_transformation_db(val):
                if val is None:
                    return None
                s = str(val).strip()
                if "%" in s and ("/" in s or "up" in s.lower()):
                    return s
                m = re.search(r'(\d{1,3})', s)
                if m:
                    num = int(m.group(1))
                    if num < 0: num = 0
                    if num > 100: num = 100
                    other = 100 - num
                    return f"{num}% up-skilling / {other}% automation"
                return None

            if "Task Transformation %" in new_tasks.columns:
                new_tasks["Task Transformation %"] = new_tasks["Task Transformation %"].apply(normalize_task_transformation_db)
            else:
                new_tasks["Task Transformation %"] = None

            try:
                # Show DB progress UI while committing manual DB update
                db_progress = st.progress(0.0)
                db_status = st.empty()

                db_status.text("Preparing data...")
                db_progress.progress(0.05)

                upsert_all_jobs_sql(engine, new_tasks)
                db_progress.progress(0.6)
                db_status.text("Tasks upserted... preparing syntheses")

                syn_rows_for_db = []
                for role, syn in st.session_state["new_synthesis"].items():
                    jd_text = st.session_state["new_jd_text"].get(role, "")
                    syn_rows_for_db.append({
                        "job_title": role,
                        "synthesis": syn,
                        "run_id": datetime.now().isoformat(timespec="seconds"),
                        "jd_hash": hashlib.sha256(jd_text.strip().encode("utf-8")).hexdigest()[:12]
                    })

                db_status.text("Inserting synthesis rows...")
                db_progress.progress(0.9)

                if syn_rows_for_db:
                    append_synthesis_sql(engine, syn_rows_for_db)

                db_progress.progress(1.0)
                db_status.text("Done")
                st.success(f"✅ Database updated with {len(new_tasks)} rows (Excel not modified).")

            except Exception as e:
                st.error(f"Failed to update Database: {e}")

            # Clear buffers after DB commit to avoid duplicate commits
            st.session_state["new_reports"].clear()
            st.session_state["new_synthesis"].clear()
            st.session_state["new_jd_text"].clear()
            st.rerun()

# -------------------------
# Dropdown Filter + Plotly Graph
# -------------------------
st.divider()
st.subheader("📊 View Generated Insights by Job Title")

if engine is not None:
    try:
        # Step 1: Fetch unique job titles
        with engine.connect() as conn:
            df_titles = pd.read_sql("SELECT DISTINCT job_title FROM all_jobs ORDER BY job_title;", conn)

        if not df_titles.empty:
            # Step 2: Dropdown for job title
            selected_title = st.selectbox(
                "Select a Job Title to Visualize Its Data:",
                df_titles['job_title'].tolist(),
                index=None,
                placeholder="Choose a job title..."
            )

            if selected_title:
                # Step 3: Query data for selected job title
                with engine.connect() as conn:
                    query_data = text("""
                        SELECT job_title, task, ai_impact_score, job_category, 
                               ai_automation_complexity, impact_explanation
                        FROM all_jobs
                        WHERE job_title = :title
                        ORDER BY task;
                    """)
                    df_selected = pd.read_sql(query_data, conn, params={"title": selected_title})

                if not df_selected.empty:
                    st.success(f"📈 Showing analytics for: **{selected_title}**")

                    # Step 4: Plotly visualization options
                    tab1, tab2, tab3 = st.tabs(["AI Impact by Task", "AI Automation Complexity", "Job Category Distribution"])

                    # --- Chart 1: Bar Chart for AI Impact Score per Task ---
                    with tab1:
                        fig1 = px.bar(
                            df_selected,
                            x="task",
                            y="ai_impact_score",
                            title=f"AI Impact Score by Task for {selected_title}",
                            text="ai_impact_score",
                            color="ai_impact_score",
                            color_continuous_scale="Blues"
                        )
                        fig1.update_layout(xaxis_tickangle=-45)
                        st.plotly_chart(fig1, use_container_width=True)

                    # --- Chart 2: Pie Chart for AI Automation Complexity ---
                    with tab2:
                        if "ai_automation_complexity" in df_selected.columns:
                            fig2 = px.pie(
                                df_selected,
                                names="ai_automation_complexity",
                                title=f"Distribution of AI Automation Complexity for {selected_title}",
                                hole=0.4
                            )
                            st.plotly_chart(fig2, use_container_width=True)
                        else:
                            st.info("No AI automation complexity data available.")

                    # --- Chart 3: Job Category Distribution ---
                    with tab3:
                        if "job_category" in df_selected.columns:
                            fig3 = px.histogram(
                                df_selected,
                                x="job_category",
                                title=f"Job Category Distribution for {selected_title}",
                                color="job_category"
                            )
                            st.plotly_chart(fig3, use_container_width=True)
                        else:
                            st.info("No job category data available.")

                else:
                    st.warning("No data found for the selected job title.")
        else:
            st.info("No job titles found in the database.")
    except Exception as e:
        st.error(f"Error fetching or visualizing data: {e}")
else:
    st.warning("Database not connected — please check your connection settings.")
